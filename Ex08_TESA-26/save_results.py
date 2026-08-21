#!/usr/bin/env python3
"""
Ex08: Генерація таблиць результатів
Читає data/results_*.json → генерує таблиці у results/
  - tables_png/  (кольорові + чорно-білі)
  - tables_csv/
  - tables_xlsx/

Використання:
  python3 ex08_tables.py              # всі формати
  python3 ex08_tables.py --fmt png    # тільки PNG
  python3 ex08_tables.py --fmt xlsx   # тільки Excel
  python3 ex08_tables.py --fmt csv    # тільки CSV
"""

import json, glob, argparse
import numpy as np
from scipy import integrate
from pathlib import Path

# ── Paths ──
EXPERIMENT_DIR = Path(__file__).resolve().parent
DATA_DIR = EXPERIMENT_DIR / "data"
RESULTS_DIR = EXPERIMENT_DIR / "results"
DS_PREFIX = ''  # set in main, e.g. 'moons_'

# ── Sparsity groups (auto-detected from data) ──
# Fallback defaults; overridden by load_results()
SPS_LOW  = []
SPS_HIGH = []

# ── Methods to exclude (removed from experiment) ──
EXCLUDE = {
    'Evo-SynFlow', 'Evo-SynFlow (Adaptive)',
    'Evo-SynFlow (SymWanda)', 'Evo-SynFlow-Taylor',
}


def load_results(sp_min=0.0, sp_max=1.0):
    """Load all results_*.json and compute aggregate metrics.
    Only includes sparsities in [sp_min, sp_max] range."""
    all_r = {}
    for f in sorted(glob.glob(str(DATA_DIR / "results_*.json"))):
        data = json.load(open(f))
        for r in data.get('results', []):
            if r['Method'] in EXCLUDE:
                continue
            sp = r['Sparsity']
            if sp < sp_min - 1e-6 or sp > sp_max + 1e-6:
                continue
            key = (r['Method'], sp)
            if key not in all_r:
                all_r[key] = []
            all_r[key].append(r)

    if not all_r:
        raise RuntimeError(f"No results found in {DATA_DIR}/results_*.json")

    sps = sorted(set(s for _, s in all_r.keys()))
    methods = sorted(set(m for m, _ in all_r.keys()))

    # Auto-detect sparsity groups from actual data
    global SPS_LOW, SPS_HIGH
    SPS_LOW = [s for s in sps if s < 0.90]
    SPS_HIGH = [s for s in sps if s >= 0.90]

    rows = []
    for m in methods:
        f1s, asps, rcus = [], [], []
        for s in sps:
            entries = all_r.get((m, s), [])
            if entries:
                f1s.append(np.mean([e['F1'] for e in entries]))
                asps.append(np.mean([e.get('Actual_Sparsity', s) for e in entries]))
                rcus.append(np.mean([e.get('Time_RCU', 0) for e in entries]))
            else:
                f1s.append(0.333); asps.append(s); rcus.append(0)

        o = np.argsort(asps)
        ausc = integrate.trapezoid(np.array(f1s)[o], x=np.array(asps)[o])
        avg_rcu = np.mean(rcus)
        dev_total = np.mean(np.abs(np.array(asps) - np.array(sps)))
        dev_under = np.mean([max(0, t - a) for t, a in zip(sps, asps)])
        ausc_adj = ausc * (1 - dev_under)

        avg_signed = np.mean(np.array(asps) - np.array(sps))
        direction = "↓" if avg_signed < -0.005 else ("↑" if avg_signed > 0.005 else "=")

        f1_map = {s: f1s[i] for i, s in enumerate(sps)}
        rows.append({
            'method': m, 'ausc': ausc, 'ausc_adj': ausc_adj,
            'dev': dev_total, 'dev_under': dev_under,
            'direction': direction, 'rcu': avg_rcu, 'f1': f1_map
        })

    rows.sort(key=lambda x: -x['ausc_adj'])
    return rows, sps


def fmt_f1(v):
    return f'{v:.2f}' if v > 0.34 else '—'


def get_summary_sps():
    """Pick 3 representative high-sparsity points from actual data."""
    summary = []
    if SPS_HIGH:
        summary.append(SPS_HIGH[0])
        summary.append(SPS_HIGH[len(SPS_HIGH)//2])
        summary.append(SPS_HIGH[-1])
    return list(dict.fromkeys(summary))


# ═══════════════════════════════════════════
#  PNG tables
# ═══════════════════════════════════════════
def generate_png(rows, sps):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.colors import LinearSegmentedColormap

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    cmap = LinearSegmentedColormap.from_list('f1', ['#ff4444', '#ffaa44', '#88cc44', '#44aa44'])

    def _make_table(cell_text, col_labels, title, filename, use_color=True):
        fig, ax = plt.subplots(figsize=(14, 10))
        ax.axis('off')
        ax.set_title(title, fontsize=13, fontweight='bold', pad=15)

        if use_color:
            cell_colors = []
            for row_data in cell_text:
                colors = []
                for j, val in enumerate(row_data):
                    if j < 2:
                        colors.append('white')
                    else:
                        try:
                            v = float(val)
                            colors.append(cmap((v - 0.33) / (0.84 - 0.33)))
                        except (ValueError, TypeError):
                            colors.append('#ffcccc')
                cell_colors.append(colors)
        else:
            cell_colors = [
                ['#f0f0f0' if i % 2 == 1 else 'white'] * len(col_labels)
                for i in range(len(cell_text))
            ]

        table = ax.table(cellText=cell_text, colLabels=col_labels,
                         cellColours=cell_colors, loc='center', cellLoc='center')
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1.0, 1.35)

        hdr_color = '#2c3e50' if use_color else '#333333'
        for j in range(len(col_labels)):
            table[0, j].set_facecolor(hdr_color)
            table[0, j].set_text_props(color='white', fontweight='bold')
        table.auto_set_column_width(list(range(len(col_labels))))

        fig.tight_layout()
        fig.savefig(str(filename), dpi=200, bbox_inches='tight', facecolor='none')
        plt.close(fig)
        print(f"  Saved: {filename}")

    # --- Single unified table: F1 at all sparsities + AUSC/Dev%/AUSCa/RCU ---
    all_sps = SPS_LOW + SPS_HIGH
    if all_sps:
        sp_min = int(all_sps[0]*100)
        sp_max = int(all_sps[-1]*100)
        cols = ['#', 'Метод'] + [f'{int(s*100)}%' for s in all_sps] + ['AUSC', 'Dev%', 'AUSCa', 'RCU']
        n_f1_cols = len(all_sps)

        cells = []
        cell_colors = []
        for i, r in enumerate(rows):
            f1_vals = [fmt_f1(r['f1'].get(s, 0.333)) for s in all_sps]
            row = [str(i+1), r['method']] + f1_vals + [
                f"{r['ausc']:.3f}", f"{r['dev']*100:.1f}",
                f"{r['ausc_adj']:.3f}", f"{r['rcu']:.0f}"]
            cells.append(row)

            # Colors
            c = ['white', 'white']
            for val_str in f1_vals:
                try:
                    v = float(val_str)
                    c.append(cmap((v - 0.33) / (0.84 - 0.33)))
                except (ValueError, TypeError):
                    c.append('#ffcccc')
            c.append('#e8f4fd')  # AUSC
            d = r['dev']
            c.append('#ffe8e8' if d > 0.02 else ('#fff3e0' if d > 0.005 else '#e8f8e8'))
            c.append('#d4edfa')  # AUSCa
            c.append('white')   # RCU
            cell_colors.append(c)

        n_cols = len(cols)
        fig, ax = plt.subplots(figsize=(max(16, 2 + n_cols * 1.1), max(8, 1 + len(rows) * 0.45)))
        ax.axis('off')
        ax.set_title(f'F1-score при спарсності {sp_min}%–{sp_max}% + зведені метрики\n'
                     f'(посортовано за AUSCa, holdout 10%)',
                     fontsize=13, fontweight='bold', pad=15)
        tbl = ax.table(cellText=cells, colLabels=cols,
                       cellColours=cell_colors, loc='center', cellLoc='center')
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(8)
        tbl.scale(1.0, 1.35)
        for j in range(n_cols):
            tbl[0, j].set_facecolor('#2c3e50')
            tbl[0, j].set_text_props(color='white', fontweight='bold', fontsize=8)
        tbl.auto_set_column_width(list(range(n_cols)))
        fig.tight_layout()
        out_path = RESULTS_DIR / f'{DS_PREFIX}table.png'
        fig.savefig(str(out_path), dpi=200, bbox_inches='tight', facecolor='none')
        plt.close(fig)
        print(f"  Saved: {out_path}")

# ═══════════════════════════════════════════
#  CSV tables
# ═══════════════════════════════════════════
def generate_csv(rows, sps):
    import csv
    all_sps = SPS_LOW + SPS_HIGH
    with open(RESULTS_DIR / f'{DS_PREFIX}table.csv', 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['#', 'Метод'] + [f'{int(s*100)}%' for s in all_sps] + ['AUSC', 'Dev%', 'AUSCa', 'RCU'])
        for i, r in enumerate(rows, 1):
            f1_vals = [fmt_f1(r['f1'].get(s, 0.333)) for s in all_sps]
            w.writerow([i, r['method']] + f1_vals + [
                f"{r['ausc']:.3f}", f"{r['dev']*100:.1f}",
                f"{r['ausc_adj']:.3f}", f"{r['rcu']:.0f}"])

    print(f"  Saved: {RESULTS_DIR / f'{DS_PREFIX}table.csv'}")


# ═══════════════════════════════════════════
#  XLSX tables
# ═══════════════════════════════════════════
def generate_xlsx(rows, sps):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor='2C3E50')
    cell_font = Font(size=10)
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    def style_sheet(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = center
                if cell.row == 1:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                else:
                    cell.font = cell_font
                    if cell.row % 2 == 0:
                        cell.fill = PatternFill('solid', fgColor='F5F5F5')
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 8)

    wb = Workbook()

    ws = wb.active; ws.title = 'Результати'
    all_sps = SPS_LOW + SPS_HIGH
    ws.append(['#', 'Метод'] + [f'{int(s*100)}%' for s in all_sps] + ['AUSC', 'Dev%', 'AUSCa', 'RCU'])
    for i, r in enumerate(rows, 1):
        f1_vals = [fmt_f1(r['f1'].get(s, 0.333)) for s in all_sps]
        ws.append([i, r['method']] + f1_vals + [
            round(r['ausc'], 3), round(r['dev']*100, 1),
            round(r['ausc_adj'], 3), round(r['rcu'])])
    style_sheet(ws)

    out = RESULTS_DIR / f'{DS_PREFIX}table.xlsx'
    wb.save(str(out))
    print(f"  Saved: {out}")


# ═══════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Ex08 Table Generator')
    parser.add_argument('--fmt', choices=['png', 'csv', 'xlsx', 'all'], default='all')
    parser.add_argument('--dataset', '-d', type=str, default='moons',
                        choices=['moons', 'circles', 'spirals', 'blobs', 'cnn', 'resnet'],
                        help='Dataset to generate tables for')
    parser.add_argument('--sp-min', type=float, default=0.0,
                        help='Minimum sparsity to display (e.g. 0.50)')
    parser.add_argument('--sp-max', type=float, default=1.0,
                        help='Maximum sparsity to display (e.g. 0.97)')
    args = parser.parse_args()

    # Flat structure: results/{dataset}_{filename}
    DATA_DIR = EXPERIMENT_DIR / 'data' if args.dataset == 'moons' else EXPERIMENT_DIR / 'data' / args.dataset
    RESULTS_DIR = EXPERIMENT_DIR / 'results'
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    DS_PREFIX = f'{args.dataset}_'

    print(f"Ex08 Table Generator ({args.dataset}, sp={args.sp_min:.0%}..{args.sp_max:.0%})")
    print("=" * 40)
    rows, sps = load_results(sp_min=args.sp_min, sp_max=args.sp_max)
    print(f"Loaded {len(rows)} methods, {len(sps)} sparsity levels\n")

    if args.fmt in ('all', 'png'):
        print("[PNG]")
        generate_png(rows, sps)
    if args.fmt in ('all', 'csv'):
        print("[CSV]")
        generate_csv(rows, sps)
    if args.fmt in ('all', 'xlsx'):
        print("[XLSX]")
        generate_xlsx(rows, sps)

    print("\nDone!")
